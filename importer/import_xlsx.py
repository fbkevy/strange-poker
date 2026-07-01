"""
One-time importer: 'Poker fund.xlsx' (Poker tab) -> normalized data.json

The Poker tab is a single zero-sum ledger. Each row is a money event where the
six player money columns sum to 0. Some rows are summaries (Outstanding, P&L)
and are used only to *validate* the import, not stored.

Output schema (data.json):
{
  "players": ["Dec","Pauli","Kev","Caoimh","Dave","Fran"],
  "config": { "chipMin":..., "chipMax":..., "winDecrement":500, ... },
  "events": [
     { "id","date","rawLabel","type","note",
       "deltas": {player: euros},
       "chips":  {player: startingStack} | null,   # main games only
       "buyins": {player: rebuyCount}   | null }    # main games only
  ]
}
"""
import json
import re
import sys
from datetime import datetime, date
from pathlib import Path

import openpyxl

PLAYERS = ["Dec", "Pauli", "Kev", "Caoimh", "Dave", "Fran"]

# Column layout (1-indexed) on the Poker tab
COL_LABEL = 1
COL_DELTA = list(range(2, 8))     # 2..7  money delta per player
COL_CHECKSUM = 8
COL_CHIPS = list(range(9, 15))    # 9..14 starting chip stack per player
COL_BUYIN = list(range(15, 21))   # 15..20 rebuy count per player
COL_NOTE = 21

# Labels of summary rows we skip (but capture for validation)
SUMMARY_PREFIXES = ("Outstanding", "P&L")


def num(v):
    """Coerce a cell to float, treating blanks/junk as None."""
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s == "":
        return None
    # e.g. "8000 +free buy" -> 8000 ; "0*" -> 0
    m = re.match(r"-?\d+(\.\d+)?", s)
    return float(m.group(0)) if m else None


def classify(label, has_chips, deltas):
    """Return an event type from the row label + shape."""
    l = label.lower()
    if "settle" in l:
        return "settle"
    if "rome" in l or "sweepstake" in l or "donation" in l or "socceraid" in l or "euro 24" in l:
        return "misc"
    if "straight flush" in l or "royal" in l:
        return "bonus"
    if l.strip() == "bet" or "putin bet" in l or "eurobet" in l:
        return "bet"
    if has_chips:
        return "main"
    if "in person" in l:
        return "in_person"
    # small-stakes side rows (#2/#3/game 2/125#/#sflush) with no chip stacks
    # are the €5 after-games (or minor side pots). Distinguish sflush bonus.
    if "sflush" in l:
        return "bonus"
    return "after"


def parse_date(label):
    """Best-effort date from the label. Returns ISO string or None."""
    # openpyxl already gave datetimes for some cells as the label; handle both.
    if isinstance(label, (datetime, date)):
        return label.date().isoformat() if isinstance(label, datetime) else label.isoformat()
    s = str(label)
    # 2023-04-15 00:00:00
    m = re.search(r"(\d{4})-(\d{2})-(\d{2})", s)
    if m:
        return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    # dd/mm/yyyy or mm/dd/yyyy or d/m (year ambiguous) -> capture what we can
    m = re.search(r"(\d{1,2})/(\d{1,2})/(\d{2,4})", s)
    if m:
        a, b, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if y < 100:
            y += 2000
        # Sheet is mostly mm/dd; keep that convention, fall back if impossible
        mo, dy = (a, b) if a <= 12 else (b, a)
        try:
            return date(y, mo, dy).isoformat()
        except ValueError:
            return None
    return None  # e.g. "7/4 (settle)" year unknown, or "bet"


def main():
    src = Path(sys.argv[1]) if len(sys.argv) > 1 else Path.home() / "Downloads" / "Poker fund.xlsx"
    out = Path(sys.argv[2]) if len(sys.argv) > 2 else Path(__file__).resolve().parents[1] / "data" / "data.json"

    wb = openpyxl.load_workbook(src, data_only=True)
    ws = wb["Poker"]

    events = []
    summary = {}  # label -> {player: value}
    bad_checksums = []

    for r in range(2, ws.max_row + 1):
        raw = ws.cell(r, COL_LABEL).value
        if raw is None or str(raw).strip() == "":
            continue
        label = str(raw).strip() if not isinstance(raw, (datetime, date)) else raw

        label_str = label.isoformat() if isinstance(label, (datetime, date)) else label
        deltas = {p: num(ws.cell(r, c).value) for p, c in zip(PLAYERS, COL_DELTA)}

        # Summary rows: store for validation, don't emit as events
        if isinstance(label_str, str) and label_str.startswith(SUMMARY_PREFIXES):
            summary[label_str] = {p: (deltas[p] or 0.0) for p in PLAYERS}
            continue

        present = {p: v for p, v in deltas.items() if v is not None}
        if not present:
            continue  # no money movement on this row

        # Checksum: the six deltas must sum to (near) zero
        s = round(sum(present.values()), 2)
        if abs(s) > 0.001:
            bad_checksums.append((r, label_str, s))

        chips_raw = {p: num(ws.cell(r, c).value) for p, c in zip(PLAYERS, COL_CHIPS)}
        has_chips = any(v is not None for v in chips_raw.values())
        etype = classify(label_str if isinstance(label_str, str) else "", has_chips, present)

        chips = None
        buyins = None
        if etype == "main":
            chips = {p: chips_raw[p] for p in PLAYERS}
            buyins = {p: num(ws.cell(r, c).value) for p, c in zip(PLAYERS, COL_BUYIN)}

        events.append({
            "id": r,  # provisional; row number is a stable-enough key for import
            "srcRow": r,
            # The sheet splits P&L vs other-debts positionally at row 43.
            "block": "game" if r >= 43 else "ledger",
            "date": parse_date(label),
            "rawLabel": label_str,
            "type": etype,
            "note": (str(ws.cell(r, COL_NOTE).value).strip()
                     if ws.cell(r, COL_NOTE).value not in (None, "") else ""),
            "deltas": {p: round(present.get(p, 0.0), 2) for p in PLAYERS},
            "chips": chips,
            "buyins": buyins,
        })

    # ---- Validation against the sheet's own summary rows ----
    def totals(pred):
        return {p: round(sum(e["deltas"][p] for e in events if pred(e)), 2) for p in PLAYERS}

    # Our clean definitions (include everything, incl. 2026):
    all_totals = totals(lambda e: True)                       # true Outstanding
    poker_pnl = totals(lambda e: e["block"] == "game")        # true poker P&L

    # Reproduce the sheet EXACTLY by mirroring its (stale) cell ranges:
    #   P&L row = SUM(43:238) ; Outstanding = P&L + SUM(2:34)
    sheet_pnl = totals(lambda e: 43 <= e["srcRow"] <= 238)
    sheet_out = totals(lambda e: (43 <= e["srcRow"] <= 238) or (2 <= e["srcRow"] <= 34))

    def match(a, b):
        return a is not None and b is not None and all(
            abs(a[p] - b[p]) < 0.001 for p in PLAYERS)

    report = {
        "events_imported": len(events),
        "bad_checksums": bad_checksums,
        "reproduces_sheet_pnl_exactly": match(sheet_pnl, summary.get("P&L")),
        "reproduces_sheet_outstanding_exactly": match(sheet_out, summary.get("Outstanding")),
        "poker_pnl_incl_2026": poker_pnl,
        "outstanding_incl_2026": all_totals,
        "sheet_pnl_row": summary.get("P&L"),
        "sheet_outstanding_row": summary.get("Outstanding"),
    }

    out.parent.mkdir(parents=True, exist_ok=True)
    data = {
        "players": PLAYERS,
        "config": {
            "chipMin": 6000, "chipMax": 9000,
            "winDecrement": 500, "secondDecrement": 250,
            "lossIncrement": 500, "lossStreakForIncrement": 3,
            "mainEntry": 20, "afterEntry": 5,
            "secondPlaceShare": 0.35,
            "straightFlushBonus": 5, "royalFlushBonus": 10,
        },
        "events": events,
    }
    out.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")

    print(json.dumps(report, indent=2, ensure_ascii=False))
    print(f"\nWrote {len(events)} events -> {out}")


if __name__ == "__main__":
    main()
